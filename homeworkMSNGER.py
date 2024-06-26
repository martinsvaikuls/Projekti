##datu manipulāciju bibliotēkas
import selenium

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service

import time

##datu uzgalabašana
import unicodedata
import csv
from openpyxl import Workbook, load_workbook

##svarīgu datu uzglabāšana
import userInformation as usrI

##datu iegūšana no tīmekļa vietnes
#"""
usrName = usrI.username
passW = usrI.password

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

url = usrI.userURL
driver.get(url)
time.sleep(1)

#ievada lietotaj vardu
find = driver.find_element(By.ID, "IDToken1")
find.send_keys(usrName)

#ievada paroli
find = driver.find_element(By.ID, "IDToken2")
find.send_keys(passW)
#input()
#"Pieteikties" -> Enter
find.send_keys('\ue007')
time.sleep(5)
#input()
find = driver.find_element(By.CLASS_NAME, "gotocal")
find.click()
time.sleep(5)
workNames = []
classTime = []
className = []

webElement = driver.find_elements(By.CLASS_NAME, "name.d-inline-block")
for i in range(len(webElement)):
    #print(webElement[i])
    workNames.append(webElement[i].text)

webElement = driver.find_elements(By.XPATH, "//div[@class='row']/div[@class='col-11']")
for i in range(len(webElement)):
    classTime.append(webElement[i].text)

webElement = driver.find_elements(By.XPATH, "//div[@class='row mt-1']/div[@class='col-11']/a")
for i in range(len(webElement)):
    className.append(webElement[i].text)
#"""
###

######
#workNames = usrI.testWorks
#className = usrI.testClasses
#classTime = usrI.testTimes
######

filePath = usrI.filePathXlsx
wb = load_workbook(filePath)
ws = wb.active
maxRows = ws.max_row
oldWorkNames = []
oldClassName = []
oldClassTime = []

##iegūt jau esošos datus 
for i in range(2,maxRows+1):
    if type(ws['B'+ str(i)].value) != type(None):
        oldWorkNames.append(str(ws['B'+ str(i)].value)) 
        oldClassName.append(str(ws['C'+ str(i)].value))
        oldClassTime.append(str(ws['D'+ str(i)].value))
    ws['B'+ str(i)] = None
    ws['C'+ str(i)] = None
    ws['D'+ str(i)] = None
    #print("a")
###
#print(oldClassName, oldClassTime, oldWorkNames)
print("Mājas darbu dati iegūti")
##uzzināt kuri dati ir jauni un kuri vairs nepastāv
#jauni dati
newName = []
newClass = []
newTime = []

notifyOldName = []
notifyOldClass = []
notifyOldTime = []

if len(workNames) > len(oldWorkNames):
    for i in range(len(workNames)):

        if workNames[i] in oldWorkNames:
            notifyOldName.append(workNames[i])
            notifyOldClass.append(className[i])
            notifyOldTime.append(classTime[i])
        else:
            newName.append(workNames[i])
            newClass.append(className[i])
            newTime.append(classTime[i])
else: 
    for i in range(len(oldWorkNames)):

        if oldWorkNames[i] in workNames:
            notifyOldName.append(oldWorkNames[i])
            notifyOldClass.append(oldClassName[i])
            notifyOldTime.append(oldClassTime[i])
        else:
            newName.append(oldWorkNames[i])
            newClass.append(oldClassName[i])
            newTime.append(oldClassTime[i])



##ievietot datus excelī
dataAmount = len(workNames)
#print("length of array containing classes", dataAmount)

for i in range(dataAmount):
    try:
        ws['B'+ str(i+2)] = str(workNames[i])
        ws['C'+ str(i+2)] = str(className[i])
        ws['D'+ str(i+2)] = str(classTime[i])
        #cellref=worksheet.cell(row=i, column=4) cellref.value=
    except:
        print("oops with the xlsx")

#wb2 = Workbook()
wb.save(filePath)
#'D:\VSCODE_Programmes\Projects\Lists\prr\data2.xlsx'
wb.close()
print("Mājas darbu dati saglabāti")
driver.close()
###another program
import pywinauto
from pywinauto import win32functions
from pywinauto.application import Application
import win32gui

from pywinauto import keyboard, mouse


##pārbaudīt vai programma jau ir atvērta
import pyautogui
import sys

pathToApp = usrI.pathToApp
application = usrI.appName

appIsOn = False
for x in pyautogui.getAllWindows():
    if application == x.title:
        appIsOn = True
timeStart = time.time()
if appIsOn == False:
    print("Telegram Tiek atvērts")
    app = Application(backend = "uia").start(pathToApp)
    while appIsOn == False:
        if time.time() - timeStart > 6.0:
            print("ErrorTimeOut")
            sys.exit(0)
        for x in pyautogui.getAllWindows():
            if application == x.title:
                appIsOn = True
            if appIsOn == True:
                break
        time.sleep(0.2)

time.sleep(1)
hwnd = win32gui.FindWindow(None, application)
win32gui.MoveWindow(hwnd, 0, 0, 960, 1040, True)

######
time.sleep(3)
pywinauto.mouse.click(coords = (128, 180))
time.sleep(1)
pywinauto.mouse.click(coords = (641, 1007))
pywinauto.keyboard.send_keys("jaunie darbi")
pywinauto.keyboard.send_keys('+{ENTER}')
##!! TypeError: 'str' object cannot be interpreted as an integer   at 138+139 if in same line
for i in range(len(newName)): 
    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}") 
    helperList = newName[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = newClass[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = newTime[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')
    pywinauto.keyboard.send_keys('+{ENTER}')
    

pywinauto.keyboard.send_keys('+{ENTER}')
pywinauto.keyboard.send_keys("vecie darbi")
pywinauto.keyboard.send_keys('+{ENTER}')

for i in range(len(notifyOldName)):
    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = notifyOldName[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = notifyOldClass[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = notifyOldTime[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')
    pywinauto.keyboard.send_keys('+{ENTER}')

pywinauto.keyboard.send_keys('{ENTER}')
######

time.sleep(1)
