#"""

import helperFunctionsDir.urls as urls
import helperFunctionsDir.helperFunctions as helperFunctions
import helperFunctionsDir.productCategories as productCategories
#"""
# import urls
# import helperFunctions
# import productCategories

"""
#for testing
import Projects.ScrapingFood.helperFunctionsDir.urls as urls
import Projects.ScrapingFood.helperFunctionsDir.helperFunctions as helperFunctions
import Projects.ScrapingFood.helperFunctionsDir.productCategories as productCategories
"""

import csv
from openpyxl import load_workbook
from openpyxl import Workbook

import datetime
import os


def saveCSV_File(website, category, linkCount, name, priceBig, key):
    
    fileName = helperFunctions.productCategories.getWebsiteFileName(website, ".csv")
    pathToThisFile = os.path.dirname(__file__)
    fullWebsiteName = helperFunctions.productCategories.getPlaceName(website)
    mutatedFilePath = pathToThisFile.rsplit("\\",1)[0] + "\\ScrapedData" +"\\" + fullWebsiteName
    fileDir = mutatedFilePath + "\\" + category + "_"+key+"_" + fileName

    print(fileDir)
    with open(fileDir, "a", newline="", encoding="utf8") as appendCSVfile:
        csvWriter = csv.writer(appendCSVfile, delimiter=";")

        for i in range(len(priceBig)):
            csvFileList = list()
            csvFileList.append(datetime.date.today())
            csvFileList.append(category)
            csvFileList.append(name[i])
            csvFileList.append(priceBig[i])
            csvWriter.writerow(csvFileList)


    appendCSVfile.close()
    return linkCount
        

def saveXLSX_File(website, category, linkCount, name, priceBig, key):
    
    ## ! checks if website xlsx file exists
    fileDotExtension = ".csv"
    doesFileExist(category, website, fileDotExtension, key)
    ###^


    maximaFileLocation = helperFunctions.productCategories.getWebsiteFileName(website, ".xlsx")
    ###Work with Excel
    wb = load_workbook(filename = maximaFileLocation)

    worksheetName = category
    if worksheetName not in wb.sheetnames:
        ws = wb.create_sheet(worksheetName)
        ws = wb[worksheetName]
        ws['A1'] = 2
        ws['A2'] = 2
    else:
        ws = wb[worksheetName]
    

    newCol, linkCount = helperFunctions.writeToEXCL_File(ws, name, priceBig, linkCount)
    DATE = 2
    try:
        cell = ws.cell(DATE, newCol)
        cell.value = datetime.date.today()
    except:
        print("Wasn't able to save date")
    wb.save(maximaFileLocation)
    wb.close()
    return linkCount
        

def doesFileExist(category, website, dotFileType, key):
    doesFolderExist(website)   
    fileName = ""

    fileName = productCategories.getWebsiteFileName(website, dotFileType)
    
    pathToThisFile = os.path.dirname(__file__)
    fullWebsiteName = helperFunctions.productCategories.getPlaceName(website)
    mutatedFilePath =  pathToThisFile+"\\ScrapedData"

    fileDir = mutatedFilePath + "\\" + category + "_"+key+"_" + fileName

    if not os.path.isfile(fileDir):
        match dotFileType:
            case ".csv":
                pass

            case ".xlsx":
                print("did not find file.. Creating...")
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "apple"
                wb.save(fileDir)
                wb.close()


def doesFolderExist(website):
    
    pathToThisFile = os.path.dirname(__file__)
    fullWebsiteName = helperFunctions.productCategories.getPlaceName(website)
    mutatedFilePath = pathToThisFile.rsplit("\\",1)[0] + "\\ScrapedData"
    for i in range(2):
        if os.path.exists(mutatedFilePath):
            print("folder exists")
            pass
        else:
            os.mkdir(mutatedFilePath) 
        
        mutatedFilePath+="\\"+fullWebsiteName



def writeToEXCL_File(ws, name, priceBig, linkCount):
### writingInfo
    
    endval = ws.max_row+1
    if linkCount < 1:
        newCol = ws.max_column+1
    else:
        newCol = ws.max_column
    start = ws['A1'].value
    end = ws['A2'].value
    for j in range(len(name)):
        isNew = True
        try:
            for i in range(start, end):
                    ##Checks if there already is such product
                    if name[j] == ws['B'+ str(i)].value:
                        cell = ws.cell(i,newCol)
                        cell.value = float(priceBig[j])
                        isNew = False
            ##Writes the new product
            if isNew:
                if newCol < 3:
                    cell = ws.cell(endval,3)
                    cell.value = float(priceBig[j])
                else:
                    cell = ws.cell(endval,newCol)
                    cell.value = float(priceBig[j])
                    
                ws["B"+str(endval)] = name[j]
                endval = endval+1
                
                                    
            
            ##Probably inefficient way of coding
        
        except:
            print("saving in excel didn't work")

    try:
        ws['A2'].value = endval
    except:
        ws['A2'] = endval
    if newCol < 3:
        newCol = 3

    return newCol

ALREADYSCRAPED = "\scrapedCategories.csv"
ALREADYSCRAPEDNEW = "\scrapedCategoriesNew.csv"

def areCategoriesScraped(website, categoryNames):
    scrapingAnswer = ""
    
    pathToThisFile = os.path.dirname(__file__)
    fullWebsiteName = helperFunctions.productCategories.getPlaceName(website)
    fileDir = pathToThisFile.rsplit("\\",1)[0] + "\\ScrapedData" + ALREADYSCRAPED
    fielDirNew = pathToThisFile.rsplit("\\",1)[0] + "\\ScrapedData" + ALREADYSCRAPEDNEW
    if (os.path.exists(fileDir)):
        with open(fileDir, "r", newline="", encoding="utf8") as readCSVfile:
            csvReader = csv.reader(readCSVfile, delimiter=";")
            for row in csvReader: 
                print(row)
                print(type(row))
                if row[0] == "":
                    break
                rowWebsite = row[0]
                category = row[1]
                writtenDay = row[2]
                today = datetime.date.today()
                writtenDATE = datetime.datetime.strptime(writtenDay, "%Y-%m-%d").date()
                prevCSVDay = datetime.timedelta(days=(writtenDATE.weekday() - 1))
                prevTodayDay = datetime.timedelta(days=(today.weekday() - 1))

                if prevTodayDay == prevCSVDay:
                    if website == rowWebsite:
                        if category in categoryNames:
                            print("category has already been scraped", category, " ", website)
                            scrapingAnswer = input("do you want to scrape again y/n")
                            scrapingAnswer = scrapingAnswer.lower()

                            while scrapingAnswer != "y" and scrapingAnswer != "n":
                                scrapingAnswer = input("do you want to scrape again y/n")
                                scrapingAnswer = scrapingAnswer.lower()

                            if scrapingAnswer == "n":
                                categoryNames.remove(category)

                    with open(fielDirNew, "a", newline="", encoding="utf8") as appendCSVfile:
                        csvWrite = csv.writer(appendCSVfile, delimiter=";")
                        csvWrite.writerow(row)
                    appendCSVfile.close()


        readCSVfile.close()

        if os.path.exists(fileDir):
            os.remove(fileDir)
            print(fileDir, " has been deleted")
            if os.path.exists(fielDirNew):
                os.rename(fielDirNew, fileDir)
            else:
                print(fielDirNew, " the new file does not exist")
        else:
            print(fileDir, " file does not exist")

    return categoryNames





def scrapedCategory(website, category):
    today = datetime.date.today()
    
    pathToThisFile = os.path.dirname(__file__)
    fullWebsiteName = helperFunctions.productCategories.getPlaceName(website)
    fileDir = pathToThisFile.rsplit("\\",1)[0] + "\\ScrapedData" + ALREADYSCRAPED

    row = [website, category, today]

    with open(fileDir, "a", newline="", encoding="utf8") as appendCSVfile:
        csvWriter = csv.writer(appendCSVfile, delimiter=";")
        csvWriter.writerow(row)


    appendCSVfile.close()
    return