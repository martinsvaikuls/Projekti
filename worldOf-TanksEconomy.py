import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service

import time

import info

from unidecode import unidecode

"""
If tank names are written correctly then there won't be any problems

"""



firstFind = "sc-2df94540-0.jZAHNc"
inputText = '//input[1]'
table = "sc-b3837b6e-0.giKwjt"

#againLAzy cuz school
###Get Info for alone
winrate = '//table/tbody/tr/td[7]'
dmg = '//table/tbody/tr/td[11]'
profite = '//table/tbody/tr/td[7]'
profitePerMinute = '//table/tbody/tr/td[9]'

tr = 2
###Get Info for multiple
winrateMul = '//table/tbody/tr['+str(tr)+']/td[7]'
dmgMul = '//table/tbody/tr['+str(tr)+']/td[11]'
profiteMul = '//table/tbody/tr['+str(tr)+']/td[7]'
profitePerMinuteMul = '//table/tbody/tr['+str(tr)+']/td[9]'



nameMultiple = '//table/tbody/tr['+ str(tr) +']/td[4]/div'
nameAlone = '//table/tbody/tr/td[4]/div'

tankEconomics = "https://tomato.gg/economics/all"
tankStats = "https://tomato.gg/tank-stats/EU/recent"


def main():
    

    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    winR = []
    dmgP = []
    prof = []
    profM = []

    allTanks = info.allTanks

    driver.get(tankStats)
    time.sleep(4)
    linkCount = 1
    winR, dmgP = getInformation(driver, winR, dmgP, allTanks, winrate, winrateMul, dmg, dmgMul, linkCount)

    driver.get(tankEconomics)
    time.sleep(4)

    

    linkCount = 2

    prof, profM = getInformation(driver, prof, profM, allTanks, profite, profiteMul, profitePerMinute, profitePerMinuteMul, linkCount)


    from openpyxl import Workbook
    from openpyxl import load_workbook
    fileLocation = '.\worldOfTanks.xlsx'

    wb = load_workbook(filename = fileLocation)
    ws = wb.active

    for i in range(len(allTanks)):
        try:
            ws["B"+str(i+2)] = allTanks[i]
            ws["C"+str(i+2)] = winR[i]
            ws["D"+str(i+2)] = dmgP[i]
            ws["E"+str(i+2)] = prof[i]
            ws["F"+str(i+2)] = profM[i]
        except:
            print("probably issue with out of bounds")
    wb.save(fileLocation)
    wb.close()

def getInformation(driver, info1, info2, allTanks, infoAlone1, infoAlone1Mul, infoAlone2, infoAlone2Mul, linkCount):
    
    a1 = driver.find_element(By.CLASS_NAME, firstFind)
    a2 = a1.find_element(By.XPATH, inputText)
    for tank in allTanks:
        notIgnoreAloneName = True
        tr = 2
        #Clear stuff pls did it me
        a2.clear()
        time.sleep(1)
        a2.send_keys(tank)
        time.sleep(0.5)
        
        infoTable = driver.find_element(By.CLASS_NAME, table)
        while tr<40:

            
            infoAlone1Mul, infoAlone2Mul = updatePlacesOfStats(tr, linkCount)
            
            nameMultiple = '//table/tbody/tr['+ str(tr) +']/td[4]/div'
            print(tr)
            time.sleep(0.1)

            
            if notIgnoreAloneName:
                name = infoTable.find_element(By.XPATH, nameAlone)
                if tank == name.text:
                    info1.append(infoTable.find_element(By.XPATH, infoAlone1).text)
                    info2.append(infoTable.find_element(By.XPATH, infoAlone2).text)
                    print("we got oneNamed")
                    break
                else:
                    name = "12434yhbfdvcxb"
                    notIgnoreAloneName = False
            else:
                try:
                    name = infoTable.find_element(By.XPATH, nameMultiple)
                    print(tank == name.text)
                    print(str(tank+"   "), str(name.text+"   "), nameMultiple)
                    if tank == name.text:
                        info1.append(infoTable.find_element(By.XPATH, infoAlone1Mul).text)
                        info2.append(infoTable.find_element(By.XPATH, infoAlone2Mul).text)
                        print("we got MultiNamed")
                        break
                except:
                    try:
                        print("multi name was not gotten: " ,tr, print(name.text))
                    except:
                        print("multi name was not gotten: " ,tr) 



            tr=tr+1

    return info1, info2

def updatePlacesOfStats(tr, linkCount):
    if linkCount == 1:
        a1 = '//table/tbody/tr['+str(tr)+']/td[7]'
        a2 = '//table/tbody/tr['+str(tr)+']/td[11]'
    else:
        a1 = '//table/tbody/tr['+str(tr)+']/td[7]'
        a2 = '//table/tbody/tr['+str(tr)+']/td[9]'

    return a1, a2

main()
